#!/usr/bin/env python3
"""
XLSX Example 03: Read and Modify Excel
Shows how to load, modify, and save existing files.
"""

from openpyxl import load_workbook

def read_and_modify():
    import os

    if not os.path.exists("input.xlsx"):
        print("[WARN] input.xlsx not found")
        print("  Place an Excel file in this directory and run:")
        print("  python 03_read_modify.py")
        return

    wb = load_workbook("input.xlsx")
    ws = wb.active

    print(f"Sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")

    for row in ws.iter_rows(max_row=5, values_only=True):
        print(row)

    wb.close()
    print("[OK] Read complete")

if __name__ == "__main__":
    read_and_modify()
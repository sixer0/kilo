#!/usr/bin/env python3
"""
XLSX Example 04: Data Validation and Formatting
Creates a spreadsheet with data validation and conditional formatting.
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

def create_with_validation():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Category'
    ws['B1'] = 'Value'
    ws['C1'] = 'Status'

    ws['A2'] = 'Product A'
    ws['B2'] = 150
    ws['C2'] = 'Active'

    ws['A3'] = 'Product B'
    ws['B3'] = 250
    ws['C3'] = 'Inactive'

    dv = DataValidation(type="list", formula1='"Active,Inactive,Pending"', allow_blank=True)
    dv.error = "Select from dropdown"
    ws.add_data_validation(dv)
    dv.add("C2:C10")

    rule = CellIsRule(operator="greaterThan", formula=[200],
                      fill=PatternFill(start_color="C6EFCE"))
    ws.conditional_formatting.add("B2:B10", rule)

    wb.save("example_04_validation.xlsx")
    print("[OK] Created: example_04_validation.xlsx")

if __name__ == "__main__":
    create_with_validation()
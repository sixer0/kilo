#!/usr/bin/env python3
"""
XLSX Example 01: Basic Spreadsheet
Creates a spreadsheet with formulas and basic formatting.
"""

from openpyxl import Workbook

def create_basic_spreadsheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws['A1'] = 'Item'
    ws['B1'] = 'Quantity'
    ws['C1'] = 'Price'
    ws['D1'] = 'Total'

    data = [('Apple', 10, 0.5), ('Banana', 20, 0.3), ('Orange', 15, 0.8)]
    for i, (name, qty, price) in enumerate(data, start=2):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = qty
        ws[f'C{i}'] = price
        ws[f'D{i}'] = f'=B{i}*C{i}'

    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save("example_01_basic.xlsx")
    print("[OK] Created: example_01_basic.xlsx")

if __name__ == "__main__":
    create_basic_spreadsheet()
#!/usr/bin/env python3
"""
XLSX Example 02: Spreadsheet with Charts
Creates a spreadsheet with embedded charts.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

def create_spreadsheet_with_chart():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Month'
    ws['B1'] = 'Sales'

    data = [('Jan', 100), ('Feb', 150), ('Mar', 200), ('Apr', 175)]
    for i, (month, sales) in enumerate(data, start=2):
        ws[f'A{i}'] = month
        ws[f'B{i}'] = sales

    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Sales"
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Month"

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=5)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "D2")

    wb.save("example_02_chart.xlsx")
    print("[OK] Created: example_02_chart.xlsx")

if __name__ == "__main__":
    create_spreadsheet_with_chart()